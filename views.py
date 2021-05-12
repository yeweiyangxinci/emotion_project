from django.shortcuts import render

# Create your views here.
# coding:utf-8
from django.http import HttpResponse
from .classifiers import DictClassifier
from .Bert.serving_predict import BertClassifier
from sklearn.metrics import accuracy_score
import jieba
import re
import json
import time
import os
import datetime
from openpyxl import load_workbook
import threading

bert = BertClassifier()

class myThread (threading.Thread):
    def __init__(self, data_arr, task_name, GPU, method):
        threading.Thread.__init__(self)
        self.data_arr = data_arr
        self.task_name = task_name
        self.GPU = GPU
        self.method = method

    def run(self):
        if self.method == "method_one":
            self.result =  bert.new_server(self.data_arr, self.task_name)
        else:
            self.result = bert.eva_model(self.data_arr, self.task_name, self.GPU)

    def get_result(self):
        try:
            return self.result  # 如果子线程不使用join方法，此处可能会报没有self.result的错误
        except Exception:
            return None

def index(request):
    """
    使用情感词典做情感分析并分词
    :param request:
    :return:
    """
    if request.method == 'POST':
	# print("coming in")
        stopword_file = "sentiment_emotion/file/chinesestop.txt"
        stopwords = [line.strip() for line in open(stopword_file, encoding="utf-8").readlines()]
        ds = DictClassifier()
        bert = BertClassifier()
        result_arr = []
        id_arr = []
        request_arr = json.loads(request.body)
        text_arr = []
        for item in request_arr:
            item_id = item.get('id')
            id_arr.append(item_id)
            item_text = item.get('msg')
            item_text = re.sub(r'//@[^：:]+:', '', item_text)
            gathertime = item.get('gathertime')
            text_arr.append(item_text)
            emotionValue = float(ds.analyse_sentence(item_text.strip()))
            if emotionValue > 0:
                emotion_type = "快乐"
            else:
                emotion_type = "低落"
            line_result = []
            sentence_depart = jieba.cut(item_text.strip())
            for word in sentence_depart:
                contain_en = bool(re.search('[a-z]', word))
                is_num = bool(re.search('.*([0-9]+).*', word))
                isspace = word.isspace()
                if contain_en or isspace:
                    continue
                if is_num:
                    continue
                    
                if word == "​":
                    continue
                if word not in stopwords:
                    if word not in line_result:
                        line_result.append(word)
            dict_item = {'id': item_id, 'emotionValue': emotionValue, 'emotion': emotion_type, 'cwsArr': line_result, 'gathertime': gathertime}
            result_arr.append(dict_item)
        polarity_arr = bert.new_server(text_arr, "polarity", id_arr)
        polar_len = len(polarity_arr)
        final_arr = []
        for idx in range(0, polar_len):
            value_item = result_arr[idx]
            bert_value = polarity_arr[idx]
            if bert_value == "0":
                bert_value = "-1"
            value_item["bertValue"] = bert_value
            final_arr.append(value_item)
        return HttpResponse(json.dumps(final_arr, ensure_ascii=False))

def getHobby(request):
    """
    获取兴趣爱好接口
    :param request:
    :return:
    """
    if request.method == 'POST':
        bert = BertClassifier()
        request_arr = json.loads(request.body)
        text_arr = []
        stopword_file = "sentiment_emotion/file/chinesestop.txt"
        stopwords = [line.strip() for line in open(stopword_file, encoding="utf-8").readlines()]
        for item in request_arr:
            item_text = item.get("msg")
            line_result = []
            sentence_depart = jieba.cut(item_text.strip())
            for word in sentence_depart:
                contain_en = bool(re.search('[a-z]', word))
                isspace = word.isspace()
                is_num = bool(re.search('.*([0-9]+).*', word))
                if contain_en or isspace:
                    continue
                if is_num:
                    continue
                if word == "​":
                    continue
                if word not in stopwords:
                    if word not in line_result:
                        line_result.append(word)
            new_line = ""
            for slice_text in line_result:
                new_line = new_line + slice_text
            text_arr.append(new_line)
        polarity_arr = bert.new_server(text_arr, "text")
        return HttpResponse(json.dumps(polarity_arr, ensure_ascii=False))

def process_yunxing(request):
    """
    单独运行兴趣爱好模型
    :param request:
    :return:
    """
    if request.method == 'POST':
        bert = BertClassifier()
        input_file = r"/DataDisk_1/shiwandong_project/emotion_project/hobby_test.txt"
        data_line = [line.strip() for line in open(input_file, encoding="utf-8").readlines()]
        print("---------------")
        print(len(data_line))
        print("---------------")
        text_arr = []
        for line in data_line:
            line_data = line.split("\t")
            id = line_data[0]
            text = line_data[1]
            sentence_depart = jieba.cut(item_text.strip())
            for word in sentence_depart:
                contain_en = bool(re.search('[a-z]', word))
                isspace = word.isspace()
                is_num = bool(re.search('.*([0-9]+).*', word))
                if contain_en or isspace:
                    continue
                if is_num:
                    continue
                if word == "​":
                    continue
                if word not in stopwords:
                    if word not in line_result:
                        line_result.append(word)
            new_line = ""
            for slice_text in line_result:
                new_line = new_line + slice_text
            text_arr.append(new_line)
            polarity_arr = bert.new_server(text_arr, "text")
            print("---------end-----------")
            print(len(polarity_arr))
            print("---------end-----------")
            output_file = r"/DataDisk_1/shiwandong_project/emotion_project/result.txt"
            target = open(output_file, "w", encoding='UTF-8')
            for line in polarity_arr:
                target.write(line + "\n")


def getPolarity(request):
    """
    获取Bert极性接口
    :param request:
    :return:
    """
    if request.method == 'POST':
        bert = BertClassifier()
        request_arr = json.loads(request.body)
        text_arr = []
        stopword_file = "sentiment_emotion/file/chinesestop.txt"
        stopwords = [line.strip() for line in open(stopword_file, encoding="utf-8").readlines()]
        id_arr = []
        for item in request_arr:
            id_arr.append(item.get("id"))
            item_text = item.get("msg")
            line_result = []
            sentence_depart = jieba.cut(item_text.strip())
            for word in sentence_depart:
                contain_en = bool(re.search('[a-z]', word))
                isspace = word.isspace()
                is_num = bool(re.search('.*([0-9]+).*', word))
                if contain_en or isspace:
                    continue
                if is_num:
                    continue
                if word == "​":
                    continue
                if word not in stopwords:
                    if word not in line_result:
                        line_result.append(word)
            new_line = ""
            for slice_text in line_result:
                new_line = new_line+slice_text
            text_arr.append(new_line)
        polarity_arr = bert.new_server(text_arr, "polarity", id_arr)
        return HttpResponse(json.dumps(polarity_arr, ensure_ascii=False))

def process_polarity(request):
    if request.method == 'POST':
        bert = BertClassifier()
        input_file = r"/DataDisk_1/shiwandong_project/emotion_project/sentiment_emotion/file/情感数据源.xlsx"
        wb = load_workbook(input_file)
        ws = wb['Sheet1']
        text_arr = []
        idx = 0
        for cell in ws['B']:
            if idx < 5000:
                pre = re.compile('>(.*?)<')
                s1 = ''.join(pre.findall(str(cell.value)))
                # print(s1)
                text_arr.append(s1)
                idx += 1
            else:
                break
        polarity_arr = bert.new_server(text_arr, "polarity")
        dict = {}
        out_file = r"/DataDisk_1/shiwandong_project/emotion_project/sentiment_emotion/file/output_file.txt"
        result_arr = []
        arr_idx = 0
        for label in polarity_arr:
            line_data = text_arr[arr_idx] + "\t" + label + "\n"
            result_arr.append(line_data)
            print(line_data)
            if label in dict:
                dict[label] += 1
            else:
                dict[label] = 0
            arr_idx += 1
        print(dict)
        with open(out_file, "w", encoding="UTF-8") as resultFile:
            print(result_arr.__len__())
            resultFile.writelines(result_arr)
        return HttpResponse(json.dumps("ok", ensure_ascii=False))

def evalModel(request):
    if request.method == 'POST':
        input_file = r'/DataDisk_1/shiwandong_project/emotion_project/sentiment_emotion/file/情感数据源.xlsx'
        wb = load_workbook(input_file)
        ws = wb['Sheet1']
        text_arr = []
        idx = 0
        for cell in ws['B']:
            if idx<5000:
                pre = re.compile('>(.*?)<')
                s1 = ''.join(pre.findall(str(cell.value)))
                print(s1)
                text_arr.append(s1)
                idx += 1
            else:
                break
        print("开始运行请求")
        print("Now：", datetime.datetime.now())
        #request_arr = json.loads(request.body)
        #input_file = r'/DataDisk_1/shiwandong_project/Bert_Base/corpus/weibo_data/weibo_test.tsv'
        #input_file = r'/DataDisk_1/shiwandong_project/Bert_Base/corpus/weibo_polar/weibo_test.txt'
        #gold_labels = []
        # text_arr = []
        # for line in input_lines:
        #     para_arr = line.split("\t")
        #     #gold_labels.append(para_arr[0])
        #     text_arr.append(para_arr[1])
        print("测试集载入结束")
        print("测试集大小", len(text_arr), "条")

        print("开始处理数据")
        print("--------------------")

        # len_idx = len(text_arr)//2
        # text_one = text_arr[0:len_idx]
        # text_two = text_arr[len_idx:]

        # thread1 = myThread(text_one, "text", "gpu_two", "method_one")
        # thread2 = myThread(text_two, "text", "gpu_one", "method_two")
        # thread1.start()
        # thread2.start()
        # thread1.join()
        # thread2.join()
        # result1 = thread1.get_result()
        # result2 = thread2.get_result()

        arr_idx = 0
        while arr_idx < 5000:
            t1 = time.time()
            polarity_arr = bert.new_server(text_arr[arr_idx:arr_idx+500], "polarity")
            arr_idx += 500
            print("处理500条过程耗费时间：", (time.time() - t1), "seconds")
        #polarity_arr = bert.new_server(text_arr, "polarity")
        #result1.extend(result2)
        #result = accuracy_score(gold_labels, result1)
        # print("--------------------")
        # print("准确率:",result)
        # print("Now：", datetime.datetime.now())
        return HttpResponse(json.dumps("ok", ensure_ascii=False))
		
